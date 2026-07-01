import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font


def json_to_excel(json_path, excel_path):

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    bold = Font(bold=True)

    for section, fields in data.items():

        # Excel sheet names cannot exceed 31 characters
        sheet_name = section[:31]

        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = "Field"
        ws["B1"] = "Value"

        ws["A1"].font = bold
        ws["B1"].font = bold

        row = 2

        for field, value in fields.items():

            ws.cell(row=row, column=1).value = field
            ws.cell(row=row, column=2).value = value

            row += 1

        # Auto-fit columns
        for column in ws.columns:

            length = 0

            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        length = max(length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column_letter].width = length + 5

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    wb.save(excel_path)

    print("Excel Created Successfully")
    print(excel_path)


if __name__ == "__main__":

    json_to_excel(
        "output/final_output.json",
        "output/final_output.xlsx"
    )